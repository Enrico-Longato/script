print("VERSIONE NUOVA IN ESECUZIONE")

"""
DCAT Metadata Generator for Excel Files

Automatically creates DCAT-compliant JSON metadata describing Excel datasets.
DCAT (Data Catalog Vocabulary) is a W3C standard for describing data catalogs.

Usage:
    python excel_to_dcat.py <path_to_excel_file>
    
Example:
    python excel_to_dcat.py "C:\path\to\imprese_fvg_01_2026.xlsx"
"""

import json
import sys
import os
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook


def get_file_metadata(file_path: str) -> dict:
    """Extract basic file metadata."""
    path = Path(file_path)
    return {
        "name": path.name,
        "size_bytes": path.stat().st_size,
        "created": datetime.fromtimestamp(path.stat().st_ctime).isoformat(),
        "modified": datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
    }


def get_excel_sheets_info(file_path: str) -> list:
    """Extract information about sheets and columns from Excel file."""
    wb = load_workbook(file_path)
    sheets_info = []
    
    for sheet_name in wb.sheetnames:
        df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=0)
        
        sheet_data = {
            "name": sheet_name,
            "columns": list(df.columns),
            "column_count": len(df.columns),
            "column_types": {col: str(df[col].dtype) for col in df.columns}
        }
        
        # Get row count
        df_full = pd.read_excel(file_path, sheet_name=sheet_name)
        sheet_data["row_count"] = len(df_full)
        
        sheets_info.append(sheet_data)
    
    return sheets_info


def generate_dcat_json(file_path: str, output_path: str = None) -> dict:
    """
    Generate DCAT-compliant JSON metadata for an Excel file.
    
    Args:
        file_path: Path to the Excel file
        output_path: Optional path to save the JSON (default: same name with .json)
    
    Returns:
        Dictionary containing DCAT metadata
    """
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    
    file_metadata = get_file_metadata(file_path)
    sheets_info = get_excel_sheets_info(file_path)
    
    # Extract base name without extension
    base_name = Path(file_path).stem
    
    # Create DCAT Catalog structure
    dcat_metadata = {
        "@context": {
            "dcat": "http://www.w3.org/ns/dcat#",
            "dctermss": "http://purl.org/dc/terms/",
            "foaf": "http://xmlns.com/foaf/0.1/",
            "vcard": "http://www.w3.org/2006/vcard/ns#"
        },
        "@type": "dcat:Catalog",
        "dctermss:title": base_name,
        "dctermss:description": f"Data catalog for {file_metadata['name']}",
        "dctermss:issued": datetime.now().isoformat(),
        "dctermss:modified": file_metadata['modified'],
        "foaf:homepage": "https://github.com/Enrico-Longato/script",
        "datasets": [
            {
                "@type": "dcat:Dataset",
                "dctermss:title": sheet_info['name'],
                "dctermss:description": f"Sheet: {sheet_info['name']} from {file_metadata['name']}",
                "dctermss:issued": file_metadata['created'],
                "dctermss:modified": file_metadata['modified'],
                "dcat:distribution": [
                    {
                        "@type": "dcat:Distribution",
                        "dctermss:title": f"{base_name} - {sheet_info['name']}",
                        "dcat:mediaType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        "dcat:accessURL": file_metadata['name'],
                    }
                ],
                "metadata": {
                    "row_count": sheet_info['row_count'],
                    "column_count": sheet_info['column_count'],
                    "columns": sheet_info['columns'],
                    "data_types": sheet_info['column_types']
                }
            }
            for sheet_info in sheets_info
        ],
        "file_info": {
            "filename": file_metadata['name'],
            "size_bytes": file_metadata['size_bytes'],
            "created": file_metadata['created'],
            "modified": file_metadata['modified'],
            "sheets_count": len(sheets_info)
        }
    }
    
    # Save to JSON file if output path is specified
    if output_path is None:
        output_path = str(Path(file_path).with_suffix('.json'))
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(dcat_metadata, f, indent=2, ensure_ascii=False)
    
    print(f"✓ DCAT metadata generated successfully!")
    print(f"✓ Saved to: {output_path}")
    
    return dcat_metadata


def main():
    """Command-line interface for the DCAT metadata generator."""
    
    # If no arguments, auto-detect from data directory
    if len(sys.argv) < 2:
        # Look for Excel files in data subdirectory
        BASE_DIR = Path(__file__).resolve().parent      # FAIR/script
        PROJECT_ROOT = BASE_DIR.parent                  # FAIR
        data_path = PROJECT_ROOT / "anagrafica" / "data"

           
        if data_path.exists():
            excel_files = list(data_path.glob("imprese_fvg_*.xlsx"))
            if excel_files:
                # Sort and pick the most recent
                excel_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
                excel_file = str(excel_files[0])
                output_file = None
                print(f"Auto-detected: {excel_file}")
            else:
                print(f"No Excel files found in {data_path}")
                sys.exit(1)
        else:
            print("Usage: python excel_to_dcat.py <path_to_excel_file> [output_json_path]")
            print("\nExample:")
            print('  python excel_to_dcat.py "imprese_fvg_01_2026.xlsx"')
            print('  python excel_to_dcat.py "data.xlsx" "output_metadata.json"')
            print(f"\nData directory not found: {data_path}")
            sys.exit(1)
    else:
        excel_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    try:
        metadata = generate_dcat_json(excel_file, output_file)
        print("\nDataset Information:")
        print(f"  File: {metadata['file_info']['filename']}")
        print(f"  Sheets: {metadata['file_info']['sheets_count']}")
        print(f"  Size: {metadata['file_info']['size_bytes']} bytes")
        
    except Exception as e:
        print(f"✗ Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
