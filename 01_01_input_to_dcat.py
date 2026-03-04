import json
import sys
import os
import warnings
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# Disable openpyxl warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def get_file_metadata(file_path: str) -> dict:
    """Extract basic file metadata."""
    path = Path(file_path)

    return {
        "name": path.name,
        "size_bytes": path.stat().st_size,
        "created": datetime.fromtimestamp(path.stat().st_ctime).isoformat(),
        "modified": datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
    }


def get_excel_sheets_info(file_path: str) -> list:
    """Extract sheet information without loading full data."""

    wb = load_workbook(file_path, read_only=True, data_only=True)

    sheets_info = []

    for sheet_name in wb.sheetnames:

        ws = wb[sheet_name]

        header = next(ws.iter_rows(max_row=1, values_only=True))
        columns = [str(c) for c in header if c is not None]

        row_count = ws.max_row - 1 if ws.max_row else 0

        sheet_data = {
            "name": sheet_name,
            "columns": columns,
            "column_count": len(columns),
            "column_types": {col: "unknown" for col in columns},
            "row_count": row_count
        }

        sheets_info.append(sheet_data)

    return sheets_info


def generate_dcat_json(file_path: str, output_path: str = None):

    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return None

    if output_path is None:
        output_path = str(Path(file_path).with_suffix(".json"))

    # Do not overwrite existing metadata
    if os.path.exists(output_path):

        print(f"ℹ DCAT metadata already exists: {output_path}")
        print("ℹ Skipping generation.")

        return None

    file_metadata = get_file_metadata(file_path)
    sheets_info = get_excel_sheets_info(file_path)

    base_name = Path(file_path).stem

    dcat_metadata = {

        "@context": {
            "dcat": "http://www.w3.org/ns/dcat#",
            "dcterms": "http://purl.org/dc/terms/",
            "foaf": "http://xmlns.com/foaf/0.1/",
            "vcard": "http://www.w3.org/2006/vcard/ns#"
        },

        "@type": "dcat:Catalog",

        "dcterms:title": base_name,
        "dcterms:description": f"Data catalog for {file_metadata['name']}",
        "dcterms:issued": datetime.now().isoformat(),
        "dcterms:modified": file_metadata["modified"],

        "foaf:homepage": "https://github.com/Enrico-Longato/script",

        "datasets": [
            {
                "@type": "dcat:Dataset",

                "dcterms:title": sheet_info["name"],
                "dcterms:description": f"Sheet: {sheet_info['name']} from {file_metadata['name']}",
                "dcterms:issued": file_metadata["created"],
                "dcterms:modified": file_metadata["modified"],

                "dcat:distribution": [
                    {
                        "@type": "dcat:Distribution",
                        "dcterms:title": f"{base_name} - {sheet_info['name']}",
                        "dcat:mediaType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        "dcat:accessURL": file_metadata["name"],
                    }
                ],

                "metadata": {
                    "row_count": sheet_info["row_count"],
                    "column_count": sheet_info["column_count"],
                    "columns": sheet_info["columns"],
                    "data_types": sheet_info["column_types"]
                }
            }

            for sheet_info in sheets_info
        ],

        "file_info": {
            "filename": file_metadata["name"],
            "size_bytes": file_metadata["size_bytes"],
            "created": file_metadata["created"],
            "modified": file_metadata["modified"],
            "sheets_count": len(sheets_info)
        }
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(dcat_metadata, f, indent=2, ensure_ascii=False)

    print("✓ DCAT metadata generated successfully")
    print(f"✓ Saved to: {output_path}")

    return dcat_metadata


def main():

    if len(sys.argv) < 2:

        BASE_DIR = Path(__file__).resolve().parent
        PROJECT_ROOT = BASE_DIR.parent

        data_path = PROJECT_ROOT / "data" / "anagrafica"

        if data_path.exists():

            excel_files = list(data_path.glob("imprese_fvg_*.xlsx"))

            if excel_files:

                excel_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)

                excel_file = str(excel_files[0])

                print(f"Auto-detected: {excel_file}")

            else:

                print(f"No Excel files found in {data_path}")
                sys.exit(0)

        else:

            print("Data directory not found")
            sys.exit(0)

    else:

        excel_file = sys.argv[1]

    generate_dcat_json(excel_file)


if __name__ == "__main__":
    main()